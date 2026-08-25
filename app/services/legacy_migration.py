"""Controlled ETL helpers. Templates carry facts; application state is calculated later."""
from __future__ import annotations
import hashlib, io, json, re, uuid
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from flask import current_app
from sqlalchemy import func
from models import db, Client, Supplier, Material, MaterialCategory, Account, MigrationRun, MigrationRow, MigrationMapping

# This matrix is deliberately tied to current SQLAlchemy models, not an old XLSX schema.
TEMPLATES = {
 'CLIENTS': {'file':'01_Clients_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Client Name*','Phone','Address','Category','Notes','Legacy Expected Due']}, 'dependency':[]},
 'SUPPLIERS': {'file':'02_Suppliers_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Supplier Name*','Phone','Address','Notes','Legacy Expected Due']}, 'dependency':[]},
 'MATERIALS': {'file':'03_Materials_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Material Name*','Category','Unit','Unit Price','Notes','Legacy Expected Stock']}, 'dependency':[]},
 'ACCOUNTS': {'file':'04_Accounts_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Account Name*','Category*','Account Type*','Opening Balance','Bank Name','Account Number','Notes','Legacy Expected Balance']}, 'dependency':[]},
 'GRN': {'file':'05_GRN_Migration_Template.xlsx','sheets': {'GRN_HEADERS':['Legacy GRN Reference*','GRN Number','Date*','Supplier*','Account','Notes'], 'GRN_ITEMS':['Legacy GRN Reference*','Material*','Quantity*','Rate','Discount','Notes']}, 'dependency':['SUPPLIERS','MATERIALS']},
 'BOOKINGS': {'file':'06_Bookings_Migration_Template.xlsx','sheets': {'BOOKINGS':['Legacy Booking Reference*','Booking Number','Date*','Client*','Notes'], 'BOOKING_ITEMS':['Legacy Booking Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['CLIENTS','MATERIALS']},
 'SALES': {'file':'07_Sales_Migration_Template.xlsx','sheets': {'SALES':['Legacy Sale Reference*','Sale/Bill Number','Date*','Client*','Legacy Booking Reference','Sale Type*','Account','Notes'], 'SALE_ITEMS':['Legacy Sale Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['CLIENTS','MATERIALS']},
 'DIRECT_SALES': {'file':'08_Direct_Sales_Migration_Template.xlsx','sheets': {'SALES':['Legacy Sale Reference*','Bill Number','Date*','Client / Walk-in Name','Account','Payment Type','Notes'], 'SALE_ITEMS':['Legacy Sale Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['MATERIALS']},
 'DELIVERIES': {'file':'09_Deliveries_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Delivery Reference*','Date*','Client','Bill Number','Material*','Quantity*','Notes']}, 'dependency':['MATERIALS']},
 'PAYMENTS': {'file':'10_Payments_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Payment Reference*','Date*','Party Type*','Party*','Amount*','Account*','Payment Type*','Reference Number','Notes']}, 'dependency':['ACCOUNTS']},
 'EXPENSES': {'file':'11_Expenses_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Expense Reference*','Date*','Account*','Amount*','Category','Reference Number','Notes']}, 'dependency':['ACCOUNTS']},
 'OPENING_BALANCES': {'file':'12_Opening_Balances_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Balance Type*','Party / Account / Material*','Amount or Quantity*','Date*','Notes','Legacy Expected Balance']}, 'dependency':[]},
}

def norm(value): return re.sub(r'\s+', ' ', str(value or '').strip()).casefold()
def clean(value): return '' if value is None else str(value).strip()
def rowhash(data): return hashlib.sha256(json.dumps(data, sort_keys=True, default=str).encode()).hexdigest()
def filehash(raw): return hashlib.sha256(raw).hexdigest()
def required(header): return header.endswith('*')
def bare(header): return header.rstrip('*').strip()

def template_workbook(kind):
    spec=TEMPLATES[kind]; wb=Workbook(); ins=wb.active; ins.title='INSTRUCTIONS'
    ins.append(['LEGACY DATA MIGRATION — '+kind]); ins.append(['Purpose','Enter cleaned historical source facts only. Calculated balances, stock and profits never overwrite this application.'])
    ins.append(['How to use','Fill only the named data sheets. Keep the header row unchanged. Use YYYY-MM-DD dates and plain numbers.'])
    ins.append(['Duplicate rule','Legacy Reference is mandatory and is permanent source traceability. Re-uploading identical rows is skipped.'])
    ins.append(['Common mistake','Do not put multiple materials or parties in one cell; use one item per row on item sheets.'])
    ins.append(['Required fields','Columns marked * are required. Resolve suggested matches yourself; the system never silently selects one.'])
    ins.column_dimensions['A'].width=26; ins.column_dimensions['B'].width=115
    for sheet, headers in spec['sheets'].items():
        ws=wb.create_sheet(sheet); ws.append(headers); ws.append(['EXAMPLE-'+kind[:3]+'-001'] + ['Example value']*(len(headers)-1)); ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); ws.column_dimensions[c.column_letter].width=max(16,min(32,len(c.value)+8))
    # Composite transaction workbooks keep their named header/item sheets; a small DATA_ENTRY index preserves the standard workbook shape.
    if 'DATA_ENTRY' not in wb.sheetnames:
        index=wb.create_sheet('DATA_ENTRY'); index.append(['Use the dedicated sheets: '+', '.join(spec['sheets'].keys())]); index.append(['Do not enter data on this index sheet.'])
    ex=wb.create_sheet('EXAMPLES'); ex.append(['Read INSTRUCTIONS first. Sample rows are illustrative and must be replaced.'])
    ref=wb.create_sheet('REFERENCE_DATA'); ref.append(['Current valid values are a guide only; they never modify uploaded facts.'])
    for title, values in [('Clients',[x.name for x in Client.query.order_by(Client.name).all()]),('Suppliers',[x.name for x in Supplier.query.order_by(Supplier.name).all()]),('Materials',[x.name for x in Material.query.order_by(Material.name).all()]),('Accounts',[x.name for x in Account.query.order_by(Account.name).all()])]:
        ref.append([title]); [ref.append([v]) for v in values]
    return wb

def _problem(sheet,row,col,message,status='INVALID',suggestions=None):
    return {'sheet':sheet,'row':row,'column':col,'problem':message,'suggested_action':suggestions or '', 'status':status}
def validate_upload(kind, raw, filename, actor=''):
    if kind not in TEMPLATES: raise ValueError('Unknown official migration template.')
    try: wb=load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc: raise ValueError('The file is not a readable XLSX workbook.') from exc
    spec=TEMPLATES[kind]; issues=[]; parsed=[]
    for sheet, headers in spec['sheets'].items():
        if sheet not in wb.sheetnames: issues.append(_problem(sheet,0,'Sheet','Required sheet is missing.')); continue
        ws=wb[sheet]; actual=[clean(c.value) for c in ws[1]]
        if actual != headers: issues.append(_problem(sheet,1,'Header','Headers must exactly match the official template.')); continue
        for excel_row, cells in enumerate(ws.iter_rows(min_row=2, values_only=True),2):
            data={bare(headers[i]): clean(cells[i] if i<len(cells) else '') for i in range(len(headers))}
            if not any(data.values()) or data.get('Legacy Reference','').startswith('EXAMPLE-'): continue
            errors=[]
            for h in headers:
                if required(h) and not data[bare(h)]: errors.append(_problem(sheet,excel_row,bare(h),'This required value is missing.'))
            for key in ('Date','Quantity','Rate','Amount','Opening Balance','Amount or Quantity','Unit Price','Discount'):
                if key in data and data[key]:
                    try:
                        if key=='Date': datetime.fromisoformat(data[key].replace('Z','+00:00'))
                        else: float(data[key])
                    except ValueError: errors.append(_problem(sheet,excel_row,key,'Use YYYY-MM-DD for dates or a plain numeric value.'))
            parsed.append((sheet,excel_row,data,errors))
    refs={};
    for sheet,n,data,errors in parsed:
        ref=data.get('Legacy Reference') or data.get('Legacy GRN Reference') or data.get('Legacy Booking Reference') or data.get('Legacy Sale Reference') or data.get('Legacy Payment Reference') or data.get('Legacy Delivery Reference') or data.get('Legacy Expense Reference')
        if ref:
            k=(sheet,norm(ref)); refs.setdefault(k,[]).append(n)
            existing=MigrationMapping.query.filter_by(template_type=kind,legacy_reference=ref).first()
            if existing: errors.append(_problem(sheet,n,'Legacy Reference','This reference was already imported as record '+str(existing.entity_id),'EXACT_DUPLICATE'))
        parsed_errors=errors; issues.extend(parsed_errors)
    run=MigrationRun(run_key=str(uuid.uuid4()),template_type=kind,filename=filename,source_hash=filehash(raw),status='VALIDATED',summary_json='{}',uploaded_by=actor)
    db.session.add(run); db.session.flush()
    counts={'READY':0,'INVALID':0,'EXACT_DUPLICATE':0,'WARNING':0,'ORPHAN':0,'BLOCKED':0}
    for sheet,n,data,errors in parsed:
        statuses=[e['status'] for e in errors]; status='INVALID' if 'INVALID' in statuses else ('EXACT_DUPLICATE' if 'EXACT_DUPLICATE' in statuses else 'READY'); counts[status]=counts.get(status,0)+1
        ref=data.get('Legacy Reference') or data.get('Legacy GRN Reference') or data.get('Legacy Booking Reference') or data.get('Legacy Sale Reference') or data.get('Legacy Payment Reference') or data.get('Legacy Delivery Reference') or data.get('Legacy Expense Reference')
        db.session.add(MigrationRow(run_id=run.id,source_sheet=sheet,source_row=n,legacy_reference=ref,row_hash=rowhash(data),status=status,data_json=json.dumps(data),error_json=json.dumps(errors)))
    counts['Total Rows']=sum(counts.values()); run.summary_json=json.dumps(counts); db.session.commit(); return run, issues

def import_run(run):
    # Only low-risk master data is presently enabled. Transaction templates remain validation/dry-run only,
    # protecting stock/ledger state until a domain-service adapter is approved for each transaction type.
    if run.template_type not in {'CLIENTS','SUPPLIERS','MATERIALS','ACCOUNTS'}: raise ValueError('This transaction template is validation-ready but import is locked pending its domain-service adapter; no records were changed.')
    created=0
    for row in run.rows:
        if row.status!='READY': continue
        d=json.loads(row.data_json); ref=row.legacy_reference
        if MigrationMapping.query.filter_by(template_type=run.template_type,legacy_reference=ref).first(): row.status='EXACT_DUPLICATE'; continue
        if run.template_type=='CLIENTS':
            obj=Client(code='MIG-'+str(uuid.uuid4())[:8].upper(),name=d['Client Name'],phone=d['Phone'] or None,address=d['Address'] or None,category=d['Category'] or 'General')
        elif run.template_type=='SUPPLIERS': obj=Supplier(name=d['Supplier Name'],phone=d['Phone'] or None,address=d['Address'] or None)
        elif run.template_type=='MATERIALS':
            cat=None
            cat_name=(d.get('Category') or '').strip()
            if cat_name:
                cat=MaterialCategory.query.filter(func.lower(MaterialCategory.name)==norm(cat_name)).first()
                if not cat:
                    cat=MaterialCategory(name=cat_name, is_active=True)
                    db.session.add(cat)
                    db.session.flush()
            obj=Material(code='MIG-'+str(uuid.uuid4())[:8].upper(),name=d['Material Name'],category_id=cat.id if cat else None,unit=d['Unit'] or 'Bags',unit_price=float(d['Unit Price'] or 0))
        else: obj=Account(name=d['Account Name'],type=d['Account Type'],account_type=d['Account Type'],category=d['Category'],balance=float(d['Opening Balance'] or 0),opening_balance=float(d['Opening Balance'] or 0),bank_name=d['Bank Name'] or None,account_number=d['Account Number'] or None)
        db.session.add(obj); db.session.flush(); row.status='IMPORTED'; row.entity_type=run.template_type[:-1].title(); row.new_entity_id=obj.id; row.imported_at=datetime.utcnow(); db.session.add(MigrationMapping(template_type=run.template_type,legacy_reference=ref,entity_type=row.entity_type,entity_id=obj.id,run_id=run.id)); created+=1
    run.status='COMPLETED'; run.imported_at=datetime.utcnow(); summary=json.loads(run.summary_json); summary['Created']=created; run.summary_json=json.dumps(summary); db.session.commit(); return created
